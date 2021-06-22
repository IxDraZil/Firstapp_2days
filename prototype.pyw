import tkinter as tk
import os
import matplotlib.pyplot as plt
from pandas import DataFrame
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook
from tkinter import filedialog, Text
from tkinter import ttk
from tkinter.constants import SEL_FIRST

global day
global expectedbalance
expectedbalance = 0
global balance
balance = 0




#sheet1 = data["Sheet1"]
#print(sheet1.cell(row = 1,column = 2).value)

#sheet2 = data["Sheet2"]
#print(sheet2.cell(row = 1 ,column = 2).value)

data = load_workbook(filename="Book.xlsx")
sheet0 = data['Sheet1']
day = sheet0.cell(row = 2,column = 4).value

sheet = data['Sheet' + str(day + 1)]



for i in range(10):
    #row i collumn 4
    expectedbalance += int(sheet.cell(row = i + 2,column = 4).value)
for i in range(10):
    if sheet.cell(row = i + 2,column = 2).value == "Paid":
        balance += int(sheet.cell(row = i + 2,column = 4).value)

global datetxt
datetxt = str(sheet.cell(row = 2, column = 5).value) +"/"+ sheet.cell(row = 2, column = 6).value + "/" + str(sheet.cell(row = 2, column = 7).value)



global data2
global df2

global sheet1
global sheet2
global sheet3
global sheet4
global sheet5
global sheet6
global sheet7
global sheet8
global sheet9
global sheet10

sheet1 = data['Sheet2']
sum = 0
for i in range(2,12):
    if sheet1.cell(row = i,column = 4).value > 0:
        sum += 1
sheet1.cell(row = 2,column = 8,value = sum)

sheet2 = data['Sheet3']
sum = 0
for i in range(2,12):
    if sheet2.cell(row = i,column = 4).value > 0:
        sum += 1
sheet2.cell(row = 2,column = 8,value = sum)

sheet3 = data['Sheet4']
sum = 0
for i in range(2,12):
    if sheet3.cell(row = i,column = 4).value > 0:
        sum += 1
sheet3.cell(row = 2,column = 8,value = sum)

sheet4 = data['Sheet5']
sum = 0
for i in range(2,12):
    if sheet4.cell(row = i,column = 4).value > 0:
        sum += 1
sheet4.cell(row = 2,column = 8,value = sum)

sheet5 = data['Sheet6']
sum = 0
for i in range(2,12):
    if sheet5.cell(row = i,column = 4).value > 0:
        sum += 1
sheet5.cell(row = 2,column = 8,value = sum)

sheet6 = data['Sheet7']
sum = 0
for i in range(2,12):
    if sheet6.cell(row = i,column = 4).value > 0:
        sum += 1
sheet6.cell(row = 2,column = 8,value = sum)

sheet7 = data['Sheet8']
sum = 0
for i in range(2,12):
    if sheet7.cell(row = i,column = 4).value > 0:
        sum += 1
sheet7.cell(row = 2,column = 8,value = sum)

sheet8 = data['Sheet9']
sum = 0
for i in range(2,12):
    if sheet8.cell(row = i,column = 4).value > 0:
        sum += 1
sheet8.cell(row = 2,column = 8,value = sum)

sheet9 = data['Sheet10']
sum = 0
for i in range(2,12):
    if sheet9.cell(row = i,column = 4).value > 0:
        sum += 1
sheet9.cell(row = 2,column = 8,value = sum)

sheet10 = data['Sheet11']
sum = 0
for i in range(2,12):
    if sheet10.cell(row = i,column = 4).value > 0:
        sum += 1
sheet10.cell(row = 2,column = 8,value = sum)









data2 = {'Day': [1,2,3,4,5,6,7,8,9,10],
         'Totalroom': [sheet1.cell(row = 2,column = 8).value,
                       sheet2.cell(row = 2,column = 8).value,
                       sheet3.cell(row = 2,column = 8).value,
                       sheet4.cell(row = 2,column = 8).value,
                       sheet5.cell(row = 2,column = 8).value,
                       sheet6.cell(row = 2,column = 8).value,
                       sheet7.cell(row = 2,column = 8).value,
                       sheet8.cell(row = 2,column = 8).value,
                       sheet9.cell(row = 2,column = 8).value,
                       sheet10.cell(row = 2,column = 8).value,]
        }
df2 = DataFrame(data2,columns=['Day','Totalroom'])





LARGEFONT =("Verdana", 35)

class tkinterApp(tk.Tk):
     
    # __init__ function for class tkinterApp
    def __init__(self, *args, **kwargs):
         
        # __init__ function for class Tk
        tk.Tk.__init__(self, *args, **kwargs)
         
        # creating a container
        container = tk.Frame(self) 
        container.pack(side = "top", fill = "both", expand = True)
  
        container.grid_rowconfigure(0, weight = 1)
        container.grid_columnconfigure(0, weight = 1)
  
        # initializing frames to an empty array
        self.frames = {} 
  
        # iterating through a tuple consisting
        # of the different page layouts
        for F in (Home, Statpaid, Statbook,Statcancle, Statsetting):
  
            frame = F(container, self)
  
            # initializing frame of that object from
            # startpage, page1, page2 respectively with
            # for loop
            self.frames[F] = frame
  
            frame.grid(row = 0, column = 0, sticky ="nsew")
  
        self.show_frame(Home)
  
    # to display the current frame passed as
    # parameter
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()
  
# first window frame startpage
    def restart(self):
            tkinterApp.destroy(self)
            os.startfile("prototype.pyw")

    def nextday(self):
        global day
        day += 1
        if day == 11:
            day = 10
        sheet0.cell(row = 2,column = 4,value = day)
        data.save('Book.xlsx')
        tkinterApp.destroy(self)
        os.startfile("prototype.pyw")

    def previousday(self):
        global day
        day -= 1
        if day == 0:
            day = 1
        sheet0.cell(row = 2,column = 4,value = day)
        data.save('Book.xlsx')
        tkinterApp.destroy(self)
        os.startfile("prototype.pyw")
        



        


class Home(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)


        
        
        print ("datetxt")

        def color(x):
            if (str(x) == "Paid"):
                return "green"
            elif x == "Booked":
                return "yellow"
            elif x == "Empty":
                return "red"

        #def daychangeup():
            #sheet.cell(row = 2, column = 5,value = 1)

            #print("The day has changed")

        #def daychangedown():
            #print("The day has changed")
        def underdevelop():
            print("underdevelopment")

        


        canvas = tk.Canvas(self, height = 720, width = 1280, bg = "#F3F6F9")
        canvas.pack()


        dashstaframe = tk.Frame(self, bg = "orange",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashstaframe.place(relwidth = 0.2, relheight = 0.05, relx = 0.02, rely = 0.12)

        dashframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.12, rely = 0.17)

        dashimgframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashimgframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.02, rely = 0.17)
        
        grayframe = tk.Frame(self, bg = "#FFFFFF")
        grayframe.place(relwidth = 0.03, relheight = 0.7438, relx = 0.11, rely = 0.173)

        graphframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        graphframe.place(relwidth = 0.7, relheight = 0.18, relx = 0.25, rely = 0.74)

        balanceframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balanceframe.place(relwidth = 0.35, relheight = 0.57, relx = 0.6, rely = 0.115)#57

        priceframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        priceframe.place(relwidth = 0.35, relheight = 0.125, relx = 0.6, rely = 0.56)

        balframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balframe.place(relwidth = 0.35, relheight = 0.08, relx = 0.6, rely = 0.11)

        paidframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        paidframe.place(relwidth = 0.13, relheight = 0.231, relx = 0.27, rely = 0.2)

        cancleframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        cancleframe.place(relwidth = 0.13, relheight = 0.231, relx = 0.27, rely = 0.45)

        bookframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        bookframe.place(relwidth = 0.13, relheight = 0.231, relx = 0.41, rely = 0.2)

        settingframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        settingframe.place(relwidth = 0.13, relheight = 0.231, relx = 0.41, rely = 0.45)

        dateframe = tk.Frame(self, bg = "#000000")
        dateframe.place(relwidth = 0.15, relheight = 0.056, relx = 0.27, rely = 0.117)

        date = tk.Label(dateframe, text = datetxt,padx = 1, pady = 1, fg = "#FFFFFF",bg = "#000000",font = ("Malgun Gothic",20))
        date.pack()

        

        #####################
        statuslabel = tk.Label(dashstaframe,text = "STATUS",fg = "#000000",bg = "orange",font =  ("Malgun Gothic",19))
        statuslabel.pack(side = "top", anchor ="nw",pady = 2,padx = 10,expand = True,fill = "x")

        imglabel1 = tk.Label(dashimgframe,text = "Room 1",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel2 = tk.Label(dashimgframe,text = "Room 2",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel2.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel3 = tk.Label(dashimgframe,text = "Room 3",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel4 = tk.Label(dashimgframe,text = "Room 4",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel4.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel5 = tk.Label(dashimgframe,text = "Room 5",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel6 = tk.Label(dashimgframe,text = "Room 6",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel6.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel7 = tk.Label(dashimgframe,text = "Room 7",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel8 = tk.Label(dashimgframe,text = "Room 8",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel8.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel9 = tk.Label(dashimgframe,text = "Room 9",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel10 = tk.Label(dashimgframe,text = "Room 10",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel10.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        ##########################

        chartlabel1 = tk.Label(dashframe,text = sheet.cell(row = 2,column = 2).value,fg = color(sheet.cell(row = 2,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel2 = tk.Label(dashframe,text = sheet.cell(row = 3,column = 2).value,fg = color(sheet.cell(row = 3,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel2.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel3 = tk.Label(dashframe,text = sheet.cell(row = 4,column = 2).value,fg = color(sheet.cell(row = 4,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel4 = tk.Label(dashframe,text = sheet.cell(row = 5,column = 2).value,fg = color(sheet.cell(row = 5,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel4.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel5 = tk.Label(dashframe,text = sheet.cell(row = 6,column = 2).value,fg = color(sheet.cell(row = 6,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel6 = tk.Label(dashframe,text = sheet.cell(row = 7,column = 2).value,fg = color(sheet.cell(row = 7,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel6.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel7 = tk.Label(dashframe,text = sheet.cell(row = 8,column = 2).value,fg = color(sheet.cell(row = 8,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel8 = tk.Label(dashframe,text = sheet.cell(row = 9,column = 2).value,fg = color(sheet.cell(row = 9,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel8.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel9 = tk.Label(dashframe,text = sheet.cell(row = 10,column = 2).value,fg = color(sheet.cell(row = 10,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel10 = tk.Label(dashframe,text = sheet.cell(row = 11,column = 2).value,fg = color(sheet.cell(row = 11,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel10.pack(side = "top",anchor = "ne",pady = 12,padx = 10)
        #################################

        dayupframe = tk.Frame(self, bg = "#FFFFFF")
        dayupframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.467, rely = 0.12)
        daydownframe = tk.Frame(self, bg = "#FFFFFF")
        daydownframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.428, rely = 0.12)

        dayupbutt = tk.Button(dayupframe, text = ">", fg = "#FFFFFF", bg = "#000000", command = controller.nextday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        dayupbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        daydownbutt = tk.Button(daydownframe, text = "<", fg = "#FFFFFF", bg = "#000000", command = controller.previousday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        daydownbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        balancelabel = tk.Label(balframe,text = "BALANCE",fg = "#000000",bg = "orange",font =  ("Arial",20))
        balancelabel.pack(side = "left", anchor ="nw",expand = True,fill = "x",pady = 10)

        chartlabel = tk.Label(graphframe,text = "Chart",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)
        

        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)

        expectbalancetxt = tk.Label(balanceintframe2,text = "Expect Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        expectbalancetxt.pack(side = "top" , anchor = "s",pady = 2)

        expectbalance = tk.Label(balanceintframe2,text = (str(expectedbalance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        expectbalance.pack(expand = True,fill = tk.BOTH)

        balanceinttxt = tk.Label(balanceintframe1,text = "Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        balanceinttxt.pack(side = "top" , anchor = "s",pady = 10)

        balanceint = tk.Label(balanceintframe1,text = (str(balance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        balanceint.pack(expand = True,fill = tk.BOTH)


        transactionframe = tk.Frame(balanceframe, bg = "#211A35")
        transactionframe.place(relwidth = 0.3, relheight = 0.58, relx = 0.62, rely = 0.175)

        transactionlabel = tk.Label(transactionframe,text = "Transaction",fg = "#FFFFFF",bg = "#211A35")
        transactionlabel.pack()

        underdev = tk.Label(transactionframe,text = "Under development",fg = "#FFFFFF",bg = "#211A35")
        underdev.pack()


        ecoprice = tk.Label(priceframe, text = "• Standard = 500",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        ecoprice.pack(side = "top", anchor = "nw",padx = 25)

        normalprice = tk.Label(priceframe, text = "• Deluxe    = 550",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        normalprice.pack(side = "top", anchor = "nw",padx = 25)

        specialprice = tk.Label(priceframe, text = "• Luxury    = 750",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        specialprice.pack(side = "top", anchor = "nw",padx = 25)


        paidbutt = tk.Button(paidframe, text = "PAID", fg = "#5EE1AA", bg = "#FFFFFF", command = lambda : controller.show_frame(Statpaid),highlightthickness = 0,bd = 0, font = ("Malgun Gothic",20) )
        paidbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        bookbutt = tk.Button(bookframe, text = "BOOK", fg = "#FFE28D", bg = "#FFFFFF", command = lambda : controller.show_frame(Statbook),highlightthickness = 0,bd = 0, font = ("Malgun Gothic",20) )
        bookbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        canclebutt = tk.Button(cancleframe, text = "CANCEL", fg = "#FF8383", bg = "#FFFFFF", command = lambda : controller.show_frame(Statcancle),highlightthickness = 0,bd = 0, font = ("Malgun Gothic",20) )
        canclebutt.pack(side = "top", expand = True,fill=tk.BOTH)

        #settingbutt = tk.Button(settingframe, text = "SETTING", fg = "#C9E7FF", bg = "#FFFFFF", command = lambda : controller.show_frame(Statsetting),highlightthickness = 0,bd = 0, font = ("Malgun Gothic",20) )
        #settingbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        settingbutt = tk.Button(settingframe, text = "SETTING", fg = "#09155A", bg = "#FFFFFF", command = underdevelop,highlightthickness = 0,bd = 0, font = ("Malgun Gothic",20) )
        settingbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        settingbuttbroadcast = tk.Button(settingframe, text = "Under development", fg = "red", bg = "#FFFFFF", command = lambda : controller.show_frame(Statsetting),highlightthickness = 0,bd = 0, font = ("Malgun Gothic",13) )
        settingbuttbroadcast.pack(side = "top", expand = True,fill=tk.BOTH)





        global data2
        global df2


        figure2 = plt.Figure(figsize=(10,0), dpi=100)
        ax2 = figure2.add_subplot(111)
        ax2.patch.set_facecolor('white')
        ax2.patch.set_alpha(1)
        line2 = FigureCanvasTkAgg(figure2, graphframe)
        line2.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH)
        df2 = df2[['Day','Totalroom']].groupby('Day').sum()
        df2.plot(kind='line', legend=True, ax=ax2, color='lightblue',marker='o', fontsize=10)




class Statpaid(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
 

        def color(x):
            if (str(x) == "Paid"):
                return "green"
            elif x == "Booked":
                return "yellow"
            elif x == "Empty":
                return "red"

        
        def checkprice(type):
            if type == "Standard":
                return (500)
            elif type == "Deluxe":
                return (550)
            elif type == "Luxury":
                return (750)

        def restart(self):
            self.destroy()
            os.startfile("prototype.pyw")
            
        def saveconfig():
            self.roomchoose = inputroom.get()
            self.typechoose = typelist.get(typelist.curselection())
            print(self.roomchoose)
            print(self.typechoose)
            savecanvas = tk.Tk()

            data = load_workbook(filename="Book.xlsx")
            sheet = data['Sheet' + str(day + 1)]

            sheet.cell(row = int(self.roomchoose) + 1,column = 2,value = "Paid")
            sheet.cell(row = int(self.roomchoose) + 1,column = 3,value = self.typechoose)
            sheet.cell(row = int(self.roomchoose) + 1,column = 4,value = checkprice(self.typechoose))

            print(sheet.cell(row = int(self.roomchoose) + 1,column = 2).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 3).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 4).value)

            

            


            def des():
                data.save('Book.xlsx')
                data.close()
                savecanvas.destroy()

            canvas = tk.Canvas(savecanvas, height = 400, width = 400, bg = "#5EE1AA")
            canvas.pack()

            askframe = tk.Frame(savecanvas, bg = "#FFFFFF")
            askframe.place(relwidth = 0.5, relheight = 0.5, relx = 0.25, rely = 0.25)

            asklabel = tk.Label(askframe,text = "Room NO." + self.roomchoose,fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
            asklabel.pack()

            asklabel2 = tk.Label(askframe,text = "Room type :" + self.typechoose,fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
            asklabel2.pack()


            askbutt = tk.Button(askframe,text = "Save",fg = "#FFFFFF",bg = "#000000",command = lambda : [des(), controller.restart()])
            askbutt.pack(side = "bottom")

            
            

        canvas = tk.Canvas(self, height = 720, width = 1280, bg = "#F3F6F9")
        canvas.pack()

        dashstaframe = tk.Frame(self, bg = "orange",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashstaframe.place(relwidth = 0.2, relheight = 0.05, relx = 0.02, rely = 0.12)

        dashframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.12, rely = 0.17)

        dashimgframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashimgframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.02, rely = 0.17)
        
        grayframe = tk.Frame(self, bg = "#FFFFFF")
        grayframe.place(relwidth = 0.03, relheight = 0.7438, relx = 0.11, rely = 0.173)

        graphframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        graphframe.place(relwidth = 0.7, relheight = 0.18, relx = 0.25, rely = 0.74)

        balanceframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balanceframe.place(relwidth = 0.35, relheight = 0.57, relx = 0.6, rely = 0.115)#57

        priceframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        priceframe.place(relwidth = 0.35, relheight = 0.125, relx = 0.6, rely = 0.56)

        balframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balframe.place(relwidth = 0.35, relheight = 0.08, relx = 0.6, rely = 0.11)

        paidconfigframe = tk.Frame(self, bg = "#5EE1AA",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        paidconfigframe.place(relwidth = 0.25, relheight = 0.48, relx = 0.3, rely = 0.2)

        dateframe = tk.Frame(self, bg = "#000000")
        dateframe.place(relwidth = 0.15, relheight = 0.056, relx = 0.27, rely = 0.117)

        date = tk.Label(dateframe, text = datetxt,padx = 1, pady = 1, fg = "#FFFFFF",bg = "#000000",font = ("Malgun Gothic",20))
        date.pack()

        

        #####################
        statuslabel = tk.Label(dashstaframe,text = "STATUS",fg = "#000000",bg = "orange",font =  ("Malgun Gothic",19))
        statuslabel.pack(side = "top", anchor ="nw",pady = 2,padx = 10,expand = True,fill = "x")

        imglabel1 = tk.Label(dashimgframe,text = "Room 1",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel2 = tk.Label(dashimgframe,text = "Room 2",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel2.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel3 = tk.Label(dashimgframe,text = "Room 3",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel4 = tk.Label(dashimgframe,text = "Room 4",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel4.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel5 = tk.Label(dashimgframe,text = "Room 5",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel6 = tk.Label(dashimgframe,text = "Room 6",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel6.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel7 = tk.Label(dashimgframe,text = "Room 7",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel8 = tk.Label(dashimgframe,text = "Room 8",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel8.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel9 = tk.Label(dashimgframe,text = "Room 9",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel10 = tk.Label(dashimgframe,text = "Room 10",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel10.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        ##########################

        chartlabel1 = tk.Label(dashframe,text = sheet.cell(row = 2,column = 2).value,fg = color(sheet.cell(row = 2,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel2 = tk.Label(dashframe,text = sheet.cell(row = 3,column = 2).value,fg = color(sheet.cell(row = 3,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel2.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel3 = tk.Label(dashframe,text = sheet.cell(row = 4,column = 2).value,fg = color(sheet.cell(row = 4,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel4 = tk.Label(dashframe,text = sheet.cell(row = 5,column = 2).value,fg = color(sheet.cell(row = 5,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel4.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel5 = tk.Label(dashframe,text = sheet.cell(row = 6,column = 2).value,fg = color(sheet.cell(row = 6,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel6 = tk.Label(dashframe,text = sheet.cell(row = 7,column = 2).value,fg = color(sheet.cell(row = 7,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel6.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel7 = tk.Label(dashframe,text = sheet.cell(row = 8,column = 2).value,fg = color(sheet.cell(row = 8,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel8 = tk.Label(dashframe,text = sheet.cell(row = 9,column = 2).value,fg = color(sheet.cell(row = 9,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel8.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel9 = tk.Label(dashframe,text = sheet.cell(row = 10,column = 2).value,fg = color(sheet.cell(row = 10,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel10 = tk.Label(dashframe,text = sheet.cell(row = 11,column = 2).value,fg = color(sheet.cell(row = 11,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel10.pack(side = "top",anchor = "ne",pady = 12,padx = 10)
        #################################

        dayupframe = tk.Frame(self, bg = "#FFFFFF")
        dayupframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.467, rely = 0.12)
        daydownframe = tk.Frame(self, bg = "#FFFFFF")
        daydownframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.428, rely = 0.12)

        dayupbutt = tk.Button(dayupframe, text = ">", fg = "#FFFFFF", bg = "#000000", command = controller.nextday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        dayupbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        daydownbutt = tk.Button(daydownframe, text = "<", fg = "#FFFFFF", bg = "#000000", command = controller.previousday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        daydownbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        balancelabel = tk.Label(balframe,text = "BALANCE",fg = "#000000",bg = "orange",font =  ("Arial",20))
        balancelabel.pack(side = "left", anchor ="nw",expand = True,fill = "x",pady = 10)

    

        chartlabel = tk.Label(graphframe,text = "Chart",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)
        

        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)

        expectbalancetxt = tk.Label(balanceintframe2,text = "Expect Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        expectbalancetxt.pack(side = "top" , anchor = "s",pady = 2)

        expectbalance = tk.Label(balanceintframe2,text = (str(expectedbalance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        expectbalance.pack(expand = True,fill = tk.BOTH)

        balanceinttxt = tk.Label(balanceintframe1,text = "Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        balanceinttxt.pack(side = "top" , anchor = "s",pady = 10)

        balanceint = tk.Label(balanceintframe1,text = (str(balance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        balanceint.pack(expand = True,fill = tk.BOTH)


        transactionframe = tk.Frame(balanceframe, bg = "#211A35")
        transactionframe.place(relwidth = 0.3, relheight = 0.58, relx = 0.62, rely = 0.175)

        transactionlabel = tk.Label(transactionframe,text = "Transaction",fg = "#FFFFFF",bg = "#211A35")
        transactionlabel.pack()

        underdev = tk.Label(transactionframe,text = "Under development",fg = "#FFFFFF",bg = "#211A35")
        underdev.pack()


        ecoprice = tk.Label(priceframe, text = "• Standard = 500",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        ecoprice.pack(side = "top", anchor = "nw",padx = 25)

        normalprice = tk.Label(priceframe, text = "• Deluxe    = 550",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        normalprice.pack(side = "top", anchor = "nw",padx = 25)

        specialprice = tk.Label(priceframe, text = "• Luxury    = 750",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        specialprice.pack(side = "top", anchor = "nw",padx = 25)


        paidroomnumber = tk.Label(paidconfigframe, text = "Room No.",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#5EE1AA",font = ("Arial Rounded MT Bold",20))
        paidroomnumber.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputroom = tk.Entry(paidconfigframe,fg = "#5EE1AA", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font =  ("Arial Rounded MT Bold",15))
        inputroom.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)

        roomtype = tk.Label(paidconfigframe, text = "Room Type",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#5EE1AA",font = ("Arial Rounded MT Bold",20))
        roomtype.pack(side="top", anchor="nw",pady = 2,padx = 5)
        
        typelist = tk.Listbox(paidconfigframe,bd = 0,highlightthickness = 0,height = 3,highlightcolor  = "#FFFFFF",font = ("Arial Rounded MT Bold",15))
        typelist.pack(pady = 10)

        typewehave = ["Standard","Deluxe","Luxury"]

        for item in typewehave:
            typelist.insert("end",item)

        balancelabel = tk.Label(balanceframe,text = "Balance",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        balancelabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)


        backbutt = tk.Button(paidconfigframe, text = "Back", fg = "#5EE1AA", bg = "#FFFFFF", command = lambda : controller.show_frame(Home),highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        backbutt.pack(side = "left",padx = 10,pady = 10)

        savebutt = tk.Button(paidconfigframe, text = "Saves", fg = "#5EE1AA", bg = "#FFFFFF", command = saveconfig,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        savebutt.pack(side = "right",padx = 10,pady = 10)

       


class Statbook(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        data = load_workbook(filename="Book.xlsx")
        sheet = data['Sheet' + str(day + 1)]


        def color(x):
            if (str(x) == "Paid"):
                return "green"
            elif x == "Booked":
                return "yellow"
            elif x == "Empty":
                return "red"

        def checkprice(type):
            if type == "Standard":
                return (500)
            elif type == "Deluxe":
                return (550)
            elif type == "Luxury":
                return (750)

            
        def saveconfig():
            self.roomchoose = inputroom.get()
            self.typechoose = typelist.get(typelist.curselection())
            print(self.roomchoose)
            print(self.typechoose)
            savecanvas = tk.Tk()

            data = load_workbook(filename="Book.xlsx")
            sheet = data['Sheet' + str(day + 1)]

            sheet.cell(row = int(self.roomchoose) + 1,column = 2,value = "Booked")
            sheet.cell(row = int(self.roomchoose) + 1,column = 3,value = self.typechoose)
            sheet.cell(row = int(self.roomchoose) + 1,column = 4,value = checkprice(self.typechoose))

            print(sheet.cell(row = int(self.roomchoose) + 1,column = 2).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 3).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 4).value)

            
            
            
            def des():
                data.save('Book.xlsx')
                data.close()
                savecanvas.destroy()

            canvas = tk.Canvas(savecanvas, height = 400, width = 400, bg = "#FFE28D")
            canvas.pack()

            askframe = tk.Frame(savecanvas, bg = "#FFFFFF")
            askframe.place(relwidth = 0.5, relheight = 0.5, relx = 0.25, rely = 0.25)

            asklabel = tk.Label(askframe,text = "Room NO." + self.roomchoose,fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
            asklabel.pack()

            asklabel2 = tk.Label(askframe,text = "Room type :" + self.typechoose,fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
            asklabel2.pack()


            askbutt = tk.Button(askframe,text = "Save",fg = "#FFFFFF",bg = "#000000",command = lambda : [des(), controller.restart()])
            askbutt.pack(side = "bottom")





        canvas = tk.Canvas(self, height = 720, width = 1280, bg = "#F3F6F9")
        canvas.pack()

        dashstaframe = tk.Frame(self, bg = "orange",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashstaframe.place(relwidth = 0.2, relheight = 0.05, relx = 0.02, rely = 0.12)

        dashframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.12, rely = 0.17)

        dashimgframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashimgframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.02, rely = 0.17)
        
        grayframe = tk.Frame(self, bg = "#FFFFFF")
        grayframe.place(relwidth = 0.03, relheight = 0.7438, relx = 0.11, rely = 0.173)

        graphframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        graphframe.place(relwidth = 0.7, relheight = 0.18, relx = 0.25, rely = 0.74)

        balanceframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balanceframe.place(relwidth = 0.35, relheight = 0.57, relx = 0.6, rely = 0.115)#57

        priceframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        priceframe.place(relwidth = 0.35, relheight = 0.125, relx = 0.6, rely = 0.56)

        balframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balframe.place(relwidth = 0.35, relheight = 0.08, relx = 0.6, rely = 0.11)

        bookconfigframe = tk.Frame(self, bg = "#FFE28D",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        bookconfigframe.place(relwidth = 0.25, relheight = 0.48, relx = 0.3, rely = 0.2)

        dateframe = tk.Frame(self, bg = "#000000")
        dateframe.place(relwidth = 0.15, relheight = 0.056, relx = 0.27, rely = 0.117)

        date = tk.Label(dateframe, text = datetxt,padx = 1, pady = 1, fg = "#FFFFFF",bg = "#000000",font = ("Malgun Gothic",20))
        date.pack()

        

        #####################
        statuslabel = tk.Label(dashstaframe,text = "STATUS",fg = "#000000",bg = "orange",font =  ("Malgun Gothic",19))
        statuslabel.pack(side = "top", anchor ="nw",pady = 2,padx = 10,expand = True,fill = "x")

        imglabel1 = tk.Label(dashimgframe,text = "Room 1",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel2 = tk.Label(dashimgframe,text = "Room 2",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel2.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel3 = tk.Label(dashimgframe,text = "Room 3",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel4 = tk.Label(dashimgframe,text = "Room 4",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel4.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel5 = tk.Label(dashimgframe,text = "Room 5",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel6 = tk.Label(dashimgframe,text = "Room 6",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel6.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel7 = tk.Label(dashimgframe,text = "Room 7",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel8 = tk.Label(dashimgframe,text = "Room 8",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel8.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel9 = tk.Label(dashimgframe,text = "Room 9",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel10 = tk.Label(dashimgframe,text = "Room 10",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel10.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        ##########################

        chartlabel1 = tk.Label(dashframe,text = sheet.cell(row = 2,column = 2).value,fg = color(sheet.cell(row = 2,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel2 = tk.Label(dashframe,text = sheet.cell(row = 3,column = 2).value,fg = color(sheet.cell(row = 3,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel2.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel3 = tk.Label(dashframe,text = sheet.cell(row = 4,column = 2).value,fg = color(sheet.cell(row = 4,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel4 = tk.Label(dashframe,text = sheet.cell(row = 5,column = 2).value,fg = color(sheet.cell(row = 5,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel4.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel5 = tk.Label(dashframe,text = sheet.cell(row = 6,column = 2).value,fg = color(sheet.cell(row = 6,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel6 = tk.Label(dashframe,text = sheet.cell(row = 7,column = 2).value,fg = color(sheet.cell(row = 7,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel6.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel7 = tk.Label(dashframe,text = sheet.cell(row = 8,column = 2).value,fg = color(sheet.cell(row = 8,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel8 = tk.Label(dashframe,text = sheet.cell(row = 9,column = 2).value,fg = color(sheet.cell(row = 9,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel8.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel9 = tk.Label(dashframe,text = sheet.cell(row = 10,column = 2).value,fg = color(sheet.cell(row = 10,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel10 = tk.Label(dashframe,text = sheet.cell(row = 11,column = 2).value,fg = color(sheet.cell(row = 11,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel10.pack(side = "top",anchor = "ne",pady = 12,padx = 10)
        #################################

        dayupframe = tk.Frame(self, bg = "#FFFFFF")
        dayupframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.467, rely = 0.12)
        daydownframe = tk.Frame(self, bg = "#FFFFFF")
        daydownframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.428, rely = 0.12)

        dayupbutt = tk.Button(dayupframe, text = ">", fg = "#FFFFFF", bg = "#000000", command = controller.nextday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        dayupbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        daydownbutt = tk.Button(daydownframe, text = "<", fg = "#FFFFFF", bg = "#000000", command = controller.previousday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        daydownbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        balancelabel = tk.Label(balframe,text = "BALANCE",fg = "#000000",bg = "orange",font =  ("Arial",20))
        balancelabel.pack(side = "left", anchor ="nw",expand = True,fill = "x",pady = 10)

    

        chartlabel = tk.Label(graphframe,text = "Chart",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)
        

        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)
        
        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)

        expectbalancetxt = tk.Label(balanceintframe2,text = "Expect Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        expectbalancetxt.pack(side = "top" , anchor = "s",pady = 2)

        expectbalance = tk.Label(balanceintframe2,text = (str(expectedbalance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        expectbalance.pack(expand = True,fill = tk.BOTH)

        balanceinttxt = tk.Label(balanceintframe1,text = "Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        balanceinttxt.pack(side = "top" , anchor = "s",pady = 10)

        balanceint = tk.Label(balanceintframe1,text = (str(balance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        balanceint.pack(expand = True,fill = tk.BOTH)


        transactionframe = tk.Frame(balanceframe, bg = "#211A35")
        transactionframe.place(relwidth = 0.3, relheight = 0.58, relx = 0.62, rely = 0.175)

        transactionlabel = tk.Label(transactionframe,text = "Transaction",fg = "#FFFFFF",bg = "#211A35")
        transactionlabel.pack()

        underdev = tk.Label(transactionframe,text = "Under development",fg = "#FFFFFF",bg = "#211A35")
        underdev.pack()


        ecoprice = tk.Label(priceframe, text = "• Standard = 500",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        ecoprice.pack(side = "top", anchor = "nw",padx = 25)

        normalprice = tk.Label(priceframe, text = "• Deluxe    = 550",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        normalprice.pack(side = "top", anchor = "nw",padx = 25)

        specialprice = tk.Label(priceframe, text = "• Luxury    = 750",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        specialprice.pack(side = "top", anchor = "nw",padx = 25)



        bookroomnumber = tk.Label(bookconfigframe, text = "Room no.",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#FFE28D",font = ("Arial Rounded MT Bold",20))
        bookroomnumber.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputroom = tk.Entry(bookconfigframe,fg = "#FFE28D", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputroom.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)
        
        roomtype = tk.Label(bookconfigframe, text = "Room Type",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#FFE28D",font = ("Arial Rounded MT Bold",20))
        roomtype.pack(side="top", anchor="nw",pady = 2,padx = 5)
        
        typelist = tk.Listbox(bookconfigframe,bd = 0,highlightthickness = 0,height = 3,highlightcolor  = "#FFFFFF",font = ("Arial Rounded MT Bold",15))
        typelist.pack(pady = 10)

        balancelabel = tk.Label(balanceframe,text = "Balance",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        balancelabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)


        typewehave = ["Standard","Deluxe","Luxury"]

        for item in typewehave:
            typelist.insert("end",item)


        backbutt = tk.Button(bookconfigframe, text = "Back", fg = "#FFE28D", bg = "#FFFFFF", command = lambda : controller.show_frame(Home),highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        backbutt.pack(side = "left",padx = 10,pady = 10)

        savebutt = tk.Button(bookconfigframe, text = "Saves", fg = "#FFE28D", bg = "#FFFFFF", command = saveconfig,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        savebutt.pack(side = "right",padx = 10,pady = 10)
    
class Statcancle(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)   
        data = load_workbook(filename="Book.xlsx")
        sheet = data['Sheet' + str(day + 1)]


        def color(x):
            if (str(x) == "Paid"):
                return "green"
            elif x == "Booked":
                return "yellow"
            elif x == "Empty":
                return "red"

        def checkprice(type):
            if type == "Standard":
                return (500)
            elif type == "Deluxe":
                return (550)
            elif type == "Luxury":
                return (750)

        
            
        def saveconfig():
            self.roomchoose = inputroom.get()
            print(self.roomchoose)
            savecanvas = tk.Tk()

            data = load_workbook(filename="Book.xlsx")
            sheet = data['Sheet' + str(day + 1)]

            sheet.cell(row = int(self.roomchoose) + 1,column = 2,value = "Empty")
            sheet.cell(row = int(self.roomchoose) + 1,column = 3,value = "None")
            sheet.cell(row = int(self.roomchoose) + 1,column = 4,value = 0)

            print(sheet.cell(row = int(self.roomchoose) + 1,column = 2).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 3).value)
            print(sheet.cell(row = int(self.roomchoose) + 1,column = 4).value)

            

            def des():
                data.save('Book.xlsx')
                data.close()
                savecanvas.destroy()

            canvas = tk.Canvas(savecanvas, height = 400, width = 400, bg = "#FF8383")
            canvas.pack()

            askframe = tk.Frame(savecanvas, bg = "#FFFFFF")
            askframe.place(relwidth = 0.5, relheight = 0.5, relx = 0.25, rely = 0.25)

            asklabel = tk.Label(askframe,text = "Room NO." + self.roomchoose,fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
            asklabel.pack()

            askbutt = tk.Button(askframe,text = "Save",fg = "#FFFFFF",bg = "#000000",command = lambda : [des(), controller.restart()])
            askbutt.pack(side = "bottom")


            

        canvas = tk.Canvas(self, height = 720, width = 1280, bg = "#F3F6F9")
        canvas.pack()

        dashstaframe = tk.Frame(self, bg = "orange",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashstaframe.place(relwidth = 0.2, relheight = 0.05, relx = 0.02, rely = 0.12)

        dashframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.12, rely = 0.17)

        dashimgframe = tk.Frame(self, bg = "#FFFFFF",highlightthickness = 1,highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        dashimgframe.place(relwidth = 0.1, relheight = 0.75, relx = 0.02, rely = 0.17)
        
        grayframe = tk.Frame(self, bg = "#FFFFFF")
        grayframe.place(relwidth = 0.03, relheight = 0.7438, relx = 0.11, rely = 0.173)

        graphframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        graphframe.place(relwidth = 0.7, relheight = 0.18, relx = 0.25, rely = 0.74)

        balanceframe = tk.Frame(self, bg = "#FFFFFF",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balanceframe.place(relwidth = 0.35, relheight = 0.57, relx = 0.6, rely = 0.115)#57

        priceframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        priceframe.place(relwidth = 0.35, relheight = 0.125, relx = 0.6, rely = 0.56)

        balframe = tk.Frame(self, bg = "orange",highlightcolor = "black",highlightbackground= "black",borderwidth=1, relief="solid")
        balframe.place(relwidth = 0.35, relheight = 0.08, relx = 0.6, rely = 0.11)

        cancleconfigframe = tk.Frame(self, bg = "#FF8383")
        cancleconfigframe.place(relwidth = 0.25, relheight = 0.48, relx = 0.3, rely = 0.2)

        dateframe = tk.Frame(self, bg = "#000000")
        dateframe.place(relwidth = 0.15, relheight = 0.056, relx = 0.27, rely = 0.117)

        date = tk.Label(dateframe, text = datetxt,padx = 1, pady = 1, fg = "#FFFFFF",bg = "#000000",font = ("Malgun Gothic",20))
        date.pack()

        

        #####################
        statuslabel = tk.Label(dashstaframe,text = "STATUS",fg = "#000000",bg = "orange",font =  ("Malgun Gothic",19))
        statuslabel.pack(side = "top", anchor ="nw",pady = 2,padx = 10,expand = True,fill = "x")

        imglabel1 = tk.Label(dashimgframe,text = "Room 1",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel2 = tk.Label(dashimgframe,text = "Room 2",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel2.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel3 = tk.Label(dashimgframe,text = "Room 3",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel4 = tk.Label(dashimgframe,text = "Room 4",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel4.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel5 = tk.Label(dashimgframe,text = "Room 5",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel6 = tk.Label(dashimgframe,text = "Room 6",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel6.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel7 = tk.Label(dashimgframe,text = "Room 7",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel8 = tk.Label(dashimgframe,text = "Room 8",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel8.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel9 = tk.Label(dashimgframe,text = "Room 9",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel10 = tk.Label(dashimgframe,text = "Room 10",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        imglabel10.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        ##########################

        chartlabel1 = tk.Label(dashframe,text = sheet.cell(row = 2,column = 2).value,fg = color(sheet.cell(row = 2,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel2 = tk.Label(dashframe,text = sheet.cell(row = 3,column = 2).value,fg = color(sheet.cell(row = 3,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel2.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel3 = tk.Label(dashframe,text = sheet.cell(row = 4,column = 2).value,fg = color(sheet.cell(row = 4,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel4 = tk.Label(dashframe,text = sheet.cell(row = 5,column = 2).value,fg = color(sheet.cell(row = 5,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel4.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel5 = tk.Label(dashframe,text = sheet.cell(row = 6,column = 2).value,fg = color(sheet.cell(row = 6,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel6 = tk.Label(dashframe,text = sheet.cell(row = 7,column = 2).value,fg = color(sheet.cell(row = 7,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel6.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel7 = tk.Label(dashframe,text = sheet.cell(row = 8,column = 2).value,fg = color(sheet.cell(row = 8,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel8 = tk.Label(dashframe,text = sheet.cell(row = 9,column = 2).value,fg = color(sheet.cell(row = 9,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel8.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel9 = tk.Label(dashframe,text = sheet.cell(row = 10,column = 2).value,fg = color(sheet.cell(row = 10,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel10 = tk.Label(dashframe,text = sheet.cell(row = 11,column = 2).value,fg = color(sheet.cell(row = 11,column = 2).value),bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel10.pack(side = "top",anchor = "ne",pady = 12,padx = 10)
        #################################

        dayupframe = tk.Frame(self, bg = "#FFFFFF")
        dayupframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.467, rely = 0.12)
        daydownframe = tk.Frame(self, bg = "#FFFFFF")
        daydownframe.place(relwidth = 0.03, relheight = 0.056, relx = 0.428, rely = 0.12)

        dayupbutt = tk.Button(dayupframe, text = ">", fg = "#FFFFFF", bg = "#000000", command = controller.nextday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        dayupbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        daydownbutt = tk.Button(daydownframe, text = "<", fg = "#FFFFFF", bg = "#000000", command = controller.previousday,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        daydownbutt.pack(side = "top", expand = True,fill=tk.BOTH)

        balancelabel = tk.Label(balframe,text = "BALANCE",fg = "#000000",bg = "orange",font =  ("Arial",20))
        balancelabel.pack(side = "left", anchor ="nw",expand = True,fill = "x",pady = 10)

    

        chartlabel = tk.Label(graphframe,text = "Chart",fg = "#000000",bg = "#FFFFFF",font =  ("Malgun Gothic",15))
        chartlabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)
        

        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)

        expectbalancetxt = tk.Label(balanceintframe2,text = "Expect Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        expectbalancetxt.pack(side = "top" , anchor = "s",pady = 2)

        expectbalance = tk.Label(balanceintframe2,text = (str(expectedbalance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        expectbalance.pack(expand = True,fill = tk.BOTH)

        balanceinttxt = tk.Label(balanceintframe1,text = "Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        balanceinttxt.pack(side = "top" , anchor = "s",pady = 10)

        balanceint = tk.Label(balanceintframe1,text = (str(balance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        balanceint.pack(expand = True,fill = tk.BOTH)


        transactionframe = tk.Frame(balanceframe, bg = "#211A35")
        transactionframe.place(relwidth = 0.3, relheight = 0.58, relx = 0.62, rely = 0.175)

        transactionlabel = tk.Label(transactionframe,text = "Transaction",fg = "#FFFFFF",bg = "#211A35")
        transactionlabel.pack()

        underdev = tk.Label(transactionframe,text = "Under development",fg = "#FFFFFF",bg = "#211A35")
        underdev.pack()


        ecoprice = tk.Label(priceframe, text = "• Standard = 500",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        ecoprice.pack(side = "top", anchor = "nw",padx = 25)

        normalprice = tk.Label(priceframe, text = "• Deluxe    = 550",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        normalprice.pack(side = "top", anchor = "nw",padx = 25)

        specialprice = tk.Label(priceframe, text = "• Luxury    = 750",padx = 1, pady = 1, fg = "#000000",bg = "orange",font = ("Malgun Gothic",12))
        specialprice.pack(side = "top", anchor = "nw",padx = 25)


        cancleroomnumber = tk.Label(cancleconfigframe, text = "Room No.",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#FF8383",font = ("Arial Rounded MT Bold",20))
        cancleroomnumber.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputroom = tk.Entry(cancleconfigframe,fg = "#FF8383", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputroom.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)


        balancelabel = tk.Label(balanceframe,text = "Balance",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        balancelabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)



        backbutt = tk.Button(cancleconfigframe, text = "Back", fg = "#FF8383", bg = "#FFFFFF", command = lambda : controller.show_frame(Home),highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        backbutt.pack(side = "left",padx = 10,pady = 10)

        savebutt = tk.Button(cancleconfigframe, text = "Saves", fg = "#FF8383", bg = "#FFFFFF", command = saveconfig,highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        savebutt.pack(side = "right",padx = 10,pady = 10)

class Statsetting(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        datasett = load_workbook(filename="Book.xlsx")
        sheet0 = datasett['Sheet1']
        sheet0 = datasett['Sheet1']
        sheet = datasett['Sheet' + str(day + 1)]

        def color(x):
            if (str(x) == "Paid"):
                return "green"
            elif x == "Booked":
                return "yellow"
            elif x == "Empty":
                return "red"


        def savesettingconfig():
                self.roomamount = inputsettingroom.get()
                self.ecoprice= inputsettingeco.get()
                self.normalprice = inputsettingnormal.get()
                self.specialprice = inputsettingspecial.get()
                sheet0.cell(row = 1,column = 2,value = self.roomamount)
                sheet0.cell(row = 2,column = 2,value = self.ecoprice)
                sheet0.cell(row = 3,column = 2,value = self.normalprice)
                sheet0.cell(row = 4,column = 2,value = self.specialprice)



                print(sheet0.cell(row = 1,column = 2).value)
                print(sheet0.cell(row = 2,column = 2).value)
                print(sheet0.cell(row = 3,column = 2).value)
                print(sheet0.cell(row = 4,column = 2).value)
                
                controller.restart()
                    

            

        canvas = tk.Canvas(self, height = 720, width = 1280, bg = "#F3F6F9")
        canvas.pack()

        dashframe = tk.Frame(self, bg = "#FFFFFF")
        dashframe.place(relwidth = 0.1, relheight = 0.8, relx = 0.12, rely = 0.12)

        dashimgframe = tk.Frame(self, bg = "#FFFFFF")
        dashimgframe.place(relwidth = 0.1, relheight = 0.8, relx = 0.02, rely = 0.12)

        graphframe = tk.Frame(self, bg = "#FFFFFF")
        graphframe.place(relwidth = 0.7, relheight = 0.18, relx = 0.25, rely = 0.74)

        balanceframe = tk.Frame(self, bg = "#FFFFFF")
        balanceframe.place(relwidth = 0.35, relheight = 0.57, relx = 0.6, rely = 0.115)

        settingconfigframe = tk.Frame(self, bg = "#C9E7FF")
        settingconfigframe.place(relwidth = 0.25, relheight = 0.48, relx = 0.3, rely = 0.2)
        

         #####################
        statuslabel = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        statuslabel.pack(side = "top", anchor ="nw",pady = 8,padx = 10)

        imglabel1 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel2 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel2.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel3 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel4 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel4.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel5 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel6 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel6.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel7 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel8 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel8.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel9 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        imglabel10 = tk.Label(dashimgframe,text = "Status",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        imglabel10.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        #########################
        chartlabel1 = tk.Label(dashframe,text = sheet.cell(row = 2,column = 2).value,fg = color(sheet.cell(row = 2,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel1.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel2 = tk.Label(dashframe,text = sheet.cell(row = 3,column = 2).value,fg = color(sheet.cell(row = 3,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel2.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel3 = tk.Label(dashframe,text = sheet.cell(row = 4,column = 2).value,fg = color(sheet.cell(row = 4,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel3.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel4 = tk.Label(dashframe,text = sheet.cell(row = 5,column = 2).value,fg = color(sheet.cell(row = 5,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel4.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel5 = tk.Label(dashframe,text = sheet.cell(row = 6,column = 2).value,fg = color(sheet.cell(row = 6,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel5.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel6 = tk.Label(dashframe,text = sheet.cell(row = 7,column = 2).value,fg = color(sheet.cell(row = 7,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel6.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel7 = tk.Label(dashframe,text = sheet.cell(row = 8,column = 2).value,fg = color(sheet.cell(row = 8,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel7.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel8 = tk.Label(dashframe,text = sheet.cell(row = 9,column = 2).value,fg = color(sheet.cell(row = 9,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel8.pack(side = "top",anchor = "ne",pady = 12,padx = 10)

        chartlabel9 = tk.Label(dashframe,text = sheet.cell(row = 10,column = 2).value,fg = color(sheet.cell(row = 10,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel9.pack(side = "top", anchor ="ne",pady = 12,padx = 10)

        chartlabel10 = tk.Label(dashframe,text = sheet.cell(row = 11,column = 2).value,fg = color(sheet.cell(row = 11,column = 2).value),bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel10.pack(side = "top",anchor = "ne",pady = 12,padx = 10)
        #################################


        balanceintframe1 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe1.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.175)

        balanceintframe2 = tk.Frame(balanceframe,bg ="#211A35")
        balanceintframe2.place(relwidth  = 0.48, relheight = 0.29,relx = 0.1, rely = 0.465)

        expectbalancetxt = tk.Label(balanceintframe2,text = "Expect Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        expectbalancetxt.pack(side = "top" , anchor = "s",pady = 2)

        expectbalance = tk.Label(balanceintframe2,text = (str(expectedbalance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        expectbalance.pack(expand = True,fill = tk.BOTH)

        balanceinttxt = tk.Label(balanceintframe1,text = "Balance",fg = "#FF9E00",bg = "#211A35",font = ("Malgun Gothic",12))
        balanceinttxt.pack(side = "top" , anchor = "s",pady = 10)

        balanceint = tk.Label(balanceintframe1,text = (str(balance) + " ฿"),fg = "#FF9E00",bg = "#211A35",font = ("Arial Rounded MT Bold",12))
        balanceint.pack(expand = True,fill = tk.BOTH)

        transactionframe = tk.Frame(balanceframe, bg = "#000000")
        transactionframe.place(relwidth = 0.28, relheight = 0.68, relx = 0.62, rely = 0.1)

        transactionlabel = tk.Label(transactionframe,text = "Transaction",fg = "#FFFFFF",bg = "#211A35")
        transactionlabel.pack()

        underdev = tk.Label(transactionframe,text = "Under development",fg = "#FFFFFF",bg = "#211A35")
        underdev.pack()

        priceframe = tk.Frame(balanceframe, bg = "#5EE1AA")
        priceframe.place(relwidth = 0.3, relheight = 0.17, relx = 0.1, rely = 0.8)

        ecoprice = tk.Label(priceframe, text = "• Standard = 500",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#5EE1AA",font = ("Arial Rounded MT Bold",12))
        ecoprice.pack(side = "top", anchor = "nw")

        normalprice = tk.Label(priceframe, text = "• Deluxe = 550",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#5EE1AA",font = ("Arial Rounded MT Bold",12))
        normalprice.pack(side = "top", anchor = "nw")

        specialprice = tk.Label(priceframe, text = "• Luxury = 750",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#5EE1AA",font = ("Arial Rounded MT Bold",12))
        specialprice.pack(side = "top", anchor = "nw")



        balancelabel = tk.Label(balanceframe,text = "Balance",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        balancelabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)

        chartlabel = tk.Label(graphframe,text = "Chart",fg = "#000000",bg = "#FFFFFF",font =  ("Arial Rounded MT Bold",15))
        chartlabel.pack(side = "left", anchor ="nw",pady = 10,padx = 10)

        settingroomnumber = tk.Label(settingconfigframe, text = "Room amount",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#C9E7FF",font = ("Arial Rounded MT Bold",15))
        settingroomnumber.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputsettingroom = tk.Entry(settingconfigframe,fg = "#C9E7FF", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputsettingroom.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)

        settingecoprize = tk.Label(settingconfigframe, text = "Eco price",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#C9E7FF",font = ("Arial Rounded MT Bold",15))
        settingecoprize.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputsettingeco = tk.Entry(settingconfigframe,fg = "#C9E7FF", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputsettingeco.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)

        settingnormalprize = tk.Label(settingconfigframe, text = "Normal price",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#C9E7FF",font = ("Arial Rounded MT Bold",15))
        settingnormalprize.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputsettingnormal = tk.Entry(settingconfigframe,fg = "#C9E7FF", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputsettingnormal.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)

        settingspecialprize = tk.Label(settingconfigframe, text = "Special price",padx = 1, pady = 1, fg = "#FFFFFF",bg = "#C9E7FF",font = ("Arial Rounded MT Bold",15))
        settingspecialprize.pack(side="top", anchor="nw",pady = 2,padx = 5)

        inputsettingspecial = tk.Entry(settingconfigframe,fg = "#C9E7FF", bg = "#FFFFFF",highlightthickness = 0,bd = 0,font = ("Arial Rounded MT Bold",15))
        inputsettingspecial.pack(side="top", anchor="nw",pady = 5,padx = 22,ipadx = 50,ipady = 2)

        savebutt = tk.Button(settingconfigframe, text = "Saves", fg = "#5EE1AA", bg = "#FFFFFF", command = lambda : [savesettingconfig() ,controller.show_frame(Home)],highlightthickness = 0,bd = 0, font = ("Arial Rounded MT Bold",20) )
        savebutt.pack(side = "right",padx = 10,pady = 10)
        
app = tkinterApp()
app.mainloop()